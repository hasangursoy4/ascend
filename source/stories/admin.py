from django.contrib import admin
from django import forms
from django.contrib import messages
from django.http import HttpResponse
from .models import Content, StoryVersion, Vocabulary, CastMember , DictionaryWord


# ============================================
# EXCEL EXPORT
# ============================================

def export_content_excel(modeladmin, request, queryset):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        modeladmin.message_user(request, '❌ openpyxl yüklü değil! pip install openpyxl', level='error')
        return

    wb = openpyxl.Workbook()

    # ---- SAYFA 1: İçerikler ----
    ws1 = wb.active
    ws1.title = "İçerikler"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F3E57")

    headers1 = ['ID', 'Başlık', 'Kategori', 'Tür', 'Yıl', 'Puan', 'Sezon', 'Süre', 'Yönetmen', 'Geliştirici', 'Stüdyo', 'Poster URL']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, content in enumerate(queryset, 2):
        ws1.append([
            content.id, content.title, content.get_category_display(),
            content.genre, content.year, float(content.rating) if content.rating else 0,
            content.season_count, content.duration, content.director,
            content.developer, content.studio, content.poster_url
        ])

    # Kolon genişlikleri
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 18

    # ---- SAYFA 2: Metinler ----
    ws2 = wb.create_sheet("Metinler")
    headers2 = ['İçerik', 'Seviye', 'İngilizce Metin', 'Türkçe Metin']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    row_idx = 2
    for content in queryset:
        for version in content.versions.all():
            ws2.append([content.title, version.level, version.text_en, version.text_tr])
            row_idx += 1

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 8
    ws2.column_dimensions['C'].width = 60
    ws2.column_dimensions['D'].width = 60

    # ---- SAYFA 3: Vocabulary ----
    ws3 = wb.create_sheet("Vocabulary")
    headers3 = ['İçerik', 'Seviye', 'Kelime', 'Tür', 'Anlam', 'Örnek Cümle']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for content in queryset:
        for version in content.versions.all():
            for vocab in version.vocabularies.all():
                ws3.append([
                    content.title, version.level,
                    vocab.word, vocab.word_type, vocab.meaning, vocab.example_sentence
                ])

    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 22

    # ---- SAYFA 4: Oyuncular ----
    ws4 = wb.create_sheet("Oyuncular")
    headers4 = ['İçerik', 'Oyuncu', 'Rol', 'Fotoğraf URL', 'Sıra']
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for content in queryset:
        for cast in content.cast_members.all():
            ws4.append([content.title, cast.name, cast.role, cast.photo_url, cast.order])

    for col in ws4.columns:
        ws4.column_dimensions[col[0].column_letter].width = 22

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ascend_export.xlsx"'
    wb.save(response)
    return response

export_content_excel.short_description = "📊 Excel'e Aktar"



class CastMemberInline(admin.TabularInline):
    model = CastMember
    extra = 2
    fields = ['order', 'name', 'role', 'photo_url']


class VocabularyInline(admin.TabularInline):
    model = Vocabulary
    extra = 2
    fields = ['word', 'word_type', 'meaning', 'example_sentence']


class StoryVersionInline(admin.StackedInline):
    model = StoryVersion
    extra = 1
    fields = ['level', 'text_en', 'text_tr']
    show_change_link = True


# ============================================
# CONTENT ADMIN — Excel yükleme burada
# ============================================

class ExcelUploadForm(forms.Form):
    excel_file = forms.FileField(
        label='Excel Dosyası (.xlsx)',
        help_text='Vocab için: word | word_type | meaning | example_sentence sütunları olmalı'
    )
    import_type = forms.ChoiceField(
        choices=[('vocab', 'Vocabulary (Kelimeler)'), ('cast', 'Oyuncular')],
        label='Ne yüklenecek?'
    )
    level = forms.ChoiceField(
        choices=[('A1','A1'),('A2','A2'),('B1','B1'),('B2','B2'),('C1','C1'),('C2','C2')],
        label='Seviye (sadece Vocabulary için)',
        required=False
    )


@admin.register(Content)
class ContentAdmin(admin.ModelAdmin):
    list_display = ['title', 'category', 'year', 'rating']
    list_filter = ['category']
    inlines = [CastMemberInline, StoryVersionInline]
    actions = [export_content_excel]

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel_view), name='content_import_excel'),
        ]
        return custom + urls

    def import_excel_view(self, request):
        from django.shortcuts import render, redirect
        try:
            import openpyxl
        except ImportError:
            self.message_user(request, '❌ pip install openpyxl', level='error')
            return redirect('..')

        if request.method == 'POST' and 'excel_file' in request.FILES:
            try:
                wb = openpyxl.load_workbook(request.FILES['excel_file'])
                added_content = 0
                added_versions = 0
                added_vocab = 0
                added_cast = 0

                CATEGORY_MAP = {
                    'dizi': 'series', 'series': 'series',
                    'film': 'movie', 'movie': 'movie',
                    'animasyon': 'animation', 'animation': 'animation',
                    'oyun': 'game', 'game': 'game',
                }

                # Sayfa 1: İçerikler
                if 'İçerikler' in wb.sheetnames:
                    ws = wb['İçerikler']
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[1]: continue
                        title = str(row[1]).strip()
                        cat_raw = str(row[2]).strip().lower() if row[2] else 'series'
                        category = CATEGORY_MAP.get(cat_raw, 'series')
                        genre = str(row[3]).strip() if row[3] else ''
                        year = str(row[4]).strip() if row[4] else ''
                        rating = float(row[5]) if row[5] else 0.0
                        season = str(row[6]).strip() if row[6] else ''
                        duration = str(row[7]).strip() if row[7] else ''
                        director = str(row[8]).strip() if row[8] else ''
                        developer = str(row[9]).strip() if row[9] else ''
                        studio = str(row[10]).strip() if row[10] else ''
                        poster = str(row[11]).strip() if row[11] else ''

                        _, created = Content.objects.get_or_create(
                            title=title,
                            defaults={
                                'category': category, 'genre': genre,
                                'year': year, 'rating': rating,
                                'season_count': season, 'duration': duration,
                                'director': director, 'developer': developer,
                                'studio': studio, 'poster_url': poster,
                            }
                        )
                        if created: added_content += 1

                # Sayfa 2: Metinler
                if 'Metinler' in wb.sheetnames:
                    ws = wb['Metinler']
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0] or not row[1]: continue
                        title = str(row[0]).strip()
                        level = str(row[1]).strip()
                        text_en = str(row[2]).strip() if row[2] else ''
                        text_tr = str(row[3]).strip() if row[3] else ''
                        try:
                            content = Content.objects.get(title=title)
                            _, created = StoryVersion.objects.get_or_create(
                                content=content, level=level,
                                defaults={'text_en': text_en, 'text_tr': text_tr}
                            )
                            if created: added_versions += 1
                        except Content.DoesNotExist:
                            pass

                # Sayfa 3: Vocabulary
                if 'Vocabulary' in wb.sheetnames:
                    ws = wb['Vocabulary']
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0] or not row[1] or not row[2]: continue
                        title = str(row[0]).strip()
                        level = str(row[1]).strip()
                        word = str(row[2]).strip()
                        word_type = str(row[3]).strip() if row[3] else ''
                        meaning = str(row[4]).strip() if row[4] else ''
                        example = str(row[5]).strip() if row[5] else ''
                        try:
                            content = Content.objects.get(title=title)
                            version = content.versions.filter(level=level).first()
                            if version:
                                _, created = Vocabulary.objects.get_or_create(
                                    version=version, word=word,
                                    defaults={'word_type': word_type, 'meaning': meaning, 'example_sentence': example}
                                )
                                if created: added_vocab += 1
                        except Content.DoesNotExist:
                            pass

                # Sayfa 4: Oyuncular
                if 'Oyuncular' in wb.sheetnames:
                    ws = wb['Oyuncular']
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0] or not row[1]: continue
                        title = str(row[0]).strip()
                        name = str(row[1]).strip()
                        role = str(row[2]).strip() if row[2] else ''
                        photo = str(row[3]).strip() if row[3] else ''
                        order = int(row[4]) if row[4] else 0
                        try:
                            content = Content.objects.get(title=title)
                            _, created = CastMember.objects.get_or_create(
                                content=content, name=name,
                                defaults={'role': role, 'photo_url': photo, 'order': order}
                            )
                            if created: added_cast += 1
                        except Content.DoesNotExist:
                            pass

                self.message_user(request,
                    f'✅ {added_content} içerik, {added_versions} metin, {added_vocab} kelime, {added_cast} oyuncu eklendi!')
                return redirect('../')

            except Exception as e:
                self.message_user(request, f'❌ Hata: {e}', level='error')

        return render(request, 'admin/stories/content/import_excel.html', {
            'title': 'Excel\'den İçerik Yükle',
            'opts': self.model._meta,
        })
        extra_context = extra_context or {}
        content = Content.objects.get(pk=object_id)

        if request.method == 'POST' and 'excel_file' in request.FILES:
            try:
                import openpyxl
            except ImportError:
                messages.error(request, '❌ openpyxl yüklü değil! Terminal: pip install openpyxl')
                return super().change_view(request, object_id, form_url, extra_context)

            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                import_type = form.cleaned_data['import_type']
                level = form.cleaned_data.get('level', 'A1')

                try:
                    wb = openpyxl.load_workbook(excel_file)
                    ws = wb.active
                    added = 0
                    skipped = 0

                    if import_type == 'vocab':
                        version = content.versions.filter(level=level).first()
                        if not version:
                            messages.error(request, f'❌ {level} seviyesi bu içerikte yok! Önce oluşturun.')
                        else:
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                if not row[0]:
                                    continue
                                word = str(row[0]).strip()
                                word_type = str(row[1]).strip() if row[1] else ''
                                meaning = str(row[2]).strip() if row[2] else ''
                                example = str(row[3]).strip() if len(row) > 3 and row[3] else ''

                                _, created = Vocabulary.objects.get_or_create(
                                    version=version,
                                    word=word,
                                    defaults={'word_type': word_type, 'meaning': meaning, 'example_sentence': example}
                                )
                                if created:
                                    added += 1
                                else:
                                    skipped += 1

                    elif import_type == 'cast':
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if not row[0]:
                                continue
                            name = str(row[0]).strip()
                            role = str(row[1]).strip() if row[1] else ''
                            photo_url = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                            order = int(row[3]) if len(row) > 3 and row[3] else 0

                            _, created = CastMember.objects.get_or_create(
                                content=content,
                                name=name,
                                defaults={'role': role, 'photo_url': photo_url, 'order': order}
                            )
                            if created:
                                added += 1
                            else:
                                skipped += 1

                    if added:
                        messages.success(request, f'✅ {added} kayıt eklendi!')
                    if skipped:
                        messages.warning(request, f'⚠️ {skipped} kayıt atlandı (zaten var).')

                except Exception as e:
                    messages.error(request, f'❌ Hata: {e}')

        extra_context['excel_form'] = ExcelUploadForm()
        extra_context['content_title'] = content.title
        return super().change_view(request, object_id, form_url, extra_context)


# ============================================
# STORYVERSION ADMIN — Hızlı metin girişi
# ============================================

class QuickVocabForm(forms.Form):
    bulk_vocab = forms.CharField(
        widget=forms.Textarea(attrs={
            'rows': 12,
            'style': 'font-family: monospace; font-size: 13px; width: 100%;',
            'placeholder': 'kelime | tür | anlam | örnek cümle\nchemistry | N. | Kimya | He is a chemistry teacher.\ndifficult | ADJ. | Zor | His life is very difficult.'
        }),
        label='Toplu Kelime Girişi',
        required=False,
        help_text='Format: kelime | tür | anlam | örnek cümle'
    )


@admin.register(StoryVersion)
class StoryVersionAdmin(admin.ModelAdmin):
    list_display = ['content', 'level']
    inlines = [VocabularyInline]

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        version = StoryVersion.objects.get(pk=object_id)

        if request.method == 'POST' and 'bulk_vocab' in request.POST:
            bulk_text = request.POST.get('bulk_vocab', '').strip()
            if bulk_text:
                added = 0
                skipped = 0
                for line in bulk_text.splitlines():
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    parts = [p.strip() for p in line.split('|')]
                    if len(parts) < 3:
                        skipped += 1
                        continue
                    word, word_type = parts[0], parts[1]
                    meaning = parts[2]
                    example = parts[3] if len(parts) > 3 else ''
                    if not word or not meaning:
                        skipped += 1
                        continue
                    _, created = Vocabulary.objects.get_or_create(
                        version=version, word=word,
                        defaults={'word_type': word_type, 'meaning': meaning, 'example_sentence': example}
                    )
                    if created:
                        added += 1
                    else:
                        skipped += 1

                if added:
                    messages.success(request, f'✅ {added} kelime eklendi!')
                if skipped:
                    messages.warning(request, f'⚠️ {skipped} kelime atlandı.')

        extra_context['quick_vocab_form'] = QuickVocabForm()
        return super().change_view(request, object_id, form_url, extra_context)


@admin.register(Vocabulary)
class VocabularyAdmin(admin.ModelAdmin):
    list_display = ['word', 'word_type', 'meaning']
    list_filter = ['version__content', 'version__level']
    search_fields = ['word', 'meaning']


@admin.register(CastMember)
class CastMemberAdmin(admin.ModelAdmin):
    list_display = ['name', 'role', 'content', 'order']
    from .models import DictionaryWord

@admin.register(DictionaryWord)
class DictionaryWordAdmin(admin.ModelAdmin):
    list_display = ['word', 'word_type', 'level', 'meaning_tr']
    list_filter = ['level', 'word_type']
    search_fields = ['word', 'meaning_tr']